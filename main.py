"""Catalog Updater - main.py"""
import os
import time
from termcolor import colored #cprint
from openpyxl import load_workbook #Workbook
from openpyxl.utils import get_column_letter #column_index_from_string
from config import INPUT_DIR, template_files, OUTPUT_PATH, COLUMN_MAPPING, HEADERS

def init():
    """ Initialize the program and load the template file."""
    print("Program Starting")
    timer = time.time()

    # Check for "Template.xlsx" and "Template.xlsm" files
    #TODO maybe remove this function and make the template a file you can select in config by name
    found_template = None
    for filename in template_files:
        if os.path.isfile(filename):
            found_template = filename
            print(f"Found template file: {found_template}")
            break

    # If a template file is found, load it
    if found_template:
        print(f"Loading {found_template} into memory.")
        wb_template = load_workbook(found_template)
    else:
        wb_template = None
        print("No template file found.")

    inbox_files = []
    for filename in os.listdir(INPUT_DIR):
        f = os.path.join(INPUT_DIR, filename)
        # checking if it is a xls (xlsx, xlsm) file
        if "xls" in f:
            inbox_files.append(f)
    if inbox_files:
        # Load the first file in the inbox_files list at index [0]
        wb_data = load_workbook(inbox_files[0])
        main(wb_data, inbox_files, wb_template, timer)
    else:
        print("No valid files found in the directory.")

def main(wb_data, inbox_files, wb_template, timer):
    """ Main function to handle data."""

    print("Catalog Updater Running")
    print(colored("You have:", "cyan"), colored(f"{len(inbox_files)} files available"))
    print(inbox_files)

    ws_template = wb_template.active
    ws_data = wb_data.active
    row_count = len(tuple(ws_data.rows))

    # Copying sheets to match import number
    for file_index in range(1, len(inbox_files)):
        new_sheet = wb_template.copy_worksheet(ws_template)
        new_sheet.title = f"Sheet{file_index + 1}"
        print(f"Initializing {new_sheet.title}")

    col_list = [0]
    def dialog(lst):
        # This function used to be a dialog box but it's now just a list in config
        lst = COLUMN_MAPPING
        print(f"Selected Columns:{lst}")
        return lst
    col_list = dialog(col_list)

    def copypaster(x, y):
        if x == "0":
            print("--Skipping Column--")
        else:
            print("Copypaster running")
            # for each cell in data sheet's column X, copy data and paste it in template cell
            for cell in range(2, row_count + 1):
                data_cell = ws_data[x + str(cell)]
                # +4 on the position of the new cell to get under the headers
                new_cell = ws_template[get_column_letter(y) + str(cell + HEADERS)]
                new_cell.value = data_cell.value

    # Take in list and switch to each column as needed to read information
    def colswitcher(lst):
        print("Column Switcher run")
        # TODO: make number of switches vary on array length?
        for i, column in enumerate(lst):
            print(f"Template Col: {i + 1} / Data Col: {column}")
            # takes in a string value at x, number at y
            # x = data sheet column, y = template sheet column
            copypaster(column, i + 1)
    colswitcher(col_list)

    # switching to next sheet and file
    # -1 from length because colswitcher has already run once by default
    idx = 0
    # _ represents files
    for _ in range(0, len(inbox_files)-1):
        idx += 1
        ws_template = wb_template["Sheet" + str(idx + 1)]
        print(f"Writing to {ws_template}")
        wb_data = load_workbook(inbox_files[idx])
        ws_data = wb_data.active
        print(f"Scanning file {inbox_files[idx]}")
        colswitcher(col_list)
    end(wb_template, timer)

def end(wb_template, timer):
    """ Save the final template file and print runtime."""
    wb_template.save(OUTPUT_PATH)
    print(colored("File exported in:", "cyan"), colored("/Outbox", "white"))

    runtime = time.time()
    print(colored("Total runtime:", "cyan"),
        colored(f"{runtime - timer} seconds", "white"))

init()
